"""RegPulse CLI — generate Verilog register banks & UVM RAL models from Excel.

Usage:
    python -m src.cli --input_excel spec.xlsx --output_dir ./output
    python -m src.cli --input_excel spec.xlsx --output_dir ./output --format rtl,uvm,c_header
    python -m src.cli --input_excel spec.xlsx --output_dir ./output --rtl_only
    python -m src.cli --input_excel spec.xlsx --output_dir ./output --template_excel
    python -m src.cli --input_excel spec.xlsx --output_dir ./output --dry_run
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import traceback

from src import __version__
from src.parser.excel_parser import ExcelParser
from src.validators.validator import Validator, ValidationError
from src.generators.rtl_generator import RtlGenerator
from src.generators.apb_wrapper_generator import ApbWrapperGenerator
from src.generators.ahb_wrapper_generator import AhbWrapperGenerator
from src.generators.axi_wrapper_generator import AxiWrapperGenerator
from src.generators.uvm_generator import UvmGenerator
from src.generators.c_header_generator import CHeaderGenerator
from src.generators.json_generator import JsonGenerator
from src.generators.html_generator import HtmlGenerator
from src.generators.markdown_generator import MarkdownGenerator

logger = logging.getLogger("regpulse")

ALL_FORMATS = ["rtl", "uvm", "c_header", "json", "html", "markdown"]
DEFAULT_FORMATS = ["rtl", "uvm", "c_header", "json", "html", "markdown"]

_GENERATORS = {
    "rtl": ("RTL (Verilog)", RtlGenerator),
    "uvm": ("UVM RAL (SystemVerilog)", UvmGenerator),
    "c_header": ("C header", CHeaderGenerator),
    "json": ("JSON IR", JsonGenerator),
    "html": ("HTML documentation", HtmlGenerator),
    "markdown": ("Markdown documentation", MarkdownGenerator),
}


def setup_logging(verbose: bool = False):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def main():
    parser = argparse.ArgumentParser(
        description="RegPulse — Generate Verilog register cores, optional bus wrappers, UVM RAL, C headers, JSON, HTML, and Markdown from Excel register specs.",
    )
    parser.add_argument("--version", action="version",
                        version=f"%(prog)s {__version__}")
    parser.add_argument("--input_excel", required=True,
                        help="Path to the input .xlsx register specification file.")
    parser.add_argument("--output_dir", required=True,
                        help="Directory where generated files will be written.")
    parser.add_argument("--block_name", default="reg_top",
                        help="Name for the register block / module (default: reg_top).")
    parser.add_argument("--base_address", default="0x0",
                        help="Base address for the register map (default: 0x0).")
    parser.add_argument("--data_width", type=int, default=32,
                        help="Register data width in bits (default: 32).")
    parser.add_argument("--format", default=None,
                        help=f"Comma-separated list of output formats: {ALL_FORMATS}. "
                             f"Default: all.")
    parser.add_argument("--rtl_only", action="store_true",
                        help="Generate only Verilog RTL (shorthand for --format rtl).")
    parser.add_argument("--uvm_only", action="store_true",
                        help="Generate only UVM RAL (shorthand for --format uvm).")
    parser.add_argument("--template_excel", action="store_true",
                        help="Generate a blank Excel template file and exit.")
    parser.add_argument("--dry_run", action="store_true",
                        help="Parse and validate only; print summary without generating files.")
    parser.add_argument("--bus", default="none", choices=["none", "apb", "ahb", "axi"],
                        help="Bus protocol wrapper to generate (default: none).")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Enable verbose (DEBUG-level) logging output.")
    args = parser.parse_args()

    setup_logging(args.verbose)

    if args.rtl_only and args.uvm_only:
        parser.error("--rtl_only and --uvm_only are mutually exclusive.")

    # Validate data_width
    if args.data_width not in (8, 16, 32, 64):
        parser.error(f"--data_width must be 8, 16, 32, or 64, got {args.data_width}.")

    # Resolve output formats
    if args.rtl_only:
        formats = ["rtl"]
    elif args.uvm_only:
        formats = ["uvm"]
    elif args.format:
        formats = [f.strip() for f in args.format.split(",")]
    else:
        formats = list(DEFAULT_FORMATS)

    for f in formats:
        if f not in ALL_FORMATS:
            parser.error(f"Unknown format '{f}'. Supported: {ALL_FORMATS}")

    # Generate template Excel if requested
    if args.template_excel:
        _generate_template(args.output_dir, args.data_width)
        return

    if not os.path.isfile(args.input_excel):
        logger.error("Input file not found: %s", args.input_excel)
        sys.exit(1)

    # Parse base_address
    try:
        base_address = int(args.base_address, 0)
    except (ValueError, TypeError):
        parser.error(f"Invalid base address: '{args.base_address}'")

    # 1. Parse
    logger.info("Parsing Excel: %s", args.input_excel)
    excel_parser = ExcelParser(args.input_excel, block_name=args.block_name,
                               data_width=args.data_width,
                               base_address=base_address)
    bank = excel_parser.parse()
    logger.info("Parsed %d register(s), address space = %d bytes.",
                bank.num_registers, bank.address_space)

    # 2. Validate
    logger.info("Validating register definitions ...")
    validator = Validator(bank)
    validator.validate()
    logger.info("Validation passed.")

    # Dry run: print summary and exit
    if args.dry_run:
        _print_summary(bank)
        return

    os.makedirs(args.output_dir, exist_ok=True)

    # 3. Generate outputs
    for fmt in formats:
        desc, gen_cls = _GENERATORS[fmt]
        logger.info("Generating %s ...", desc)
        gen = gen_cls(bank)
        path = gen.generate(args.output_dir)
        logger.info("Written: %s", path)

    # 4. Optional bus wrapper
    if args.bus == "apb":
        logger.info("Generating APB4 wrapper ...")
        gen = ApbWrapperGenerator(bank)
        path = gen.generate(args.output_dir)
        logger.info("Written: %s", path)
    elif args.bus == "ahb":
        logger.info("Generating AHB-Lite wrapper ...")
        gen = AhbWrapperGenerator(bank)
        path = gen.generate(args.output_dir)
        logger.info("Written: %s", path)
    elif args.bus == "axi":
        logger.info("Generating AXI4-Lite wrapper ...")
        gen = AxiWrapperGenerator(bank)
        path = gen.generate(args.output_dir)
        logger.info("Written: %s", path)

    logger.info("Done.")


def _print_summary(bank):
    """Print a register map summary to stdout."""
    print(f"\n{'=' * 60}")
    print(f"  RegPulse Register Summary — {bank.name}")
    print(f"{'=' * 60}")
    print(f"  Base Address : 0x{bank.base_address:08X}")
    print(f"  Data Width   : {bank.data_width} bits")
    print(f"  Registers    : {bank.num_registers}")
    print(f"  Address Space: {bank.address_space} bytes")
    print(f"{'=' * 60}")
    for reg in bank.registers:
        field_info = ", ".join(
            f"{f.name}[{f.msb}:{f.lsb}]({f.access_type})" for f in reg.fields
        )
        print(f"  0x{reg.offset:03X}  {reg.name:<12s} {reg.effective_access}  "
              f"reset=0x{reg.reset_val:08X}  {field_info}")
    if bank.interrupt_pairs:
        print(f"\n  Interrupt pairs: {len(bank.interrupt_pairs)}")
        for src_f, src_r, en_f, en_r in bank.interrupt_pairs:
            print(f"    {src_f.name} ({src_r.name}) -> {en_f.name} ({en_r.name})")
    print(f"{'=' * 60}\n")


def _generate_template(output_dir: str, data_width: int = 32):
    """Generate a blank Excel template with data validation."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "register_template.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Registers"

    headers = ["Name", "Offset", "Field", "Bits", "Access", "Reset",
               "Hardware Trigger", "Side Effect", "Interrupt"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Data validation for Access column
    dv_access = DataValidation(
        type="list",
        formula1='"RW,RO,W1C,RC,RS,WO,W1S,W0C"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Access Type",
        error="Must be one of: RW, RO, W1C, RC, RS, WO, W1S, W0C"
    )
    dv_access.add(f"E2:E1000")
    ws.add_data_validation(dv_access)

    # Data validation for Hardware Trigger column
    dv_hw = DataValidation(
        type="list",
        formula1='"input,output,"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid HW Interface",
        error='Must be "input", "output", or left blank.'
    )
    dv_hw.add(f"G2:G1000")
    ws.add_data_validation(dv_hw)

    # Data validation for Interrupt column
    dv_irq = DataValidation(
        type="list",
        formula1='"source,enable,"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Interrupt Role",
        error='Must be "source", "enable", or left blank.'
    )
    dv_irq.add(f"I2:I1000")
    ws.add_data_validation(dv_irq)

    # Set column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["I"].width = 12

    # Add a sample row
    ws.append(["MY_REG", "0x00", "ENABLE", "0", "RW", "0", "output", "", ""])

    wb.save(out_path)
    logger.info("Template written: %s", out_path)


if __name__ == "__main__":
    try:
        main()
    except ValueError as exc:
        logger.error("Parse error: %s", exc)
        sys.exit(1)
    except ValidationError as exc:
        logger.error("Validation error: %s", exc)
        sys.exit(2)
    except FileNotFoundError as exc:
        logger.error("File not found: %s", exc)
        sys.exit(3)
    except Exception as exc:
        logger.critical("Unexpected error: %s", exc)
        traceback.print_exc()
        sys.exit(99)
