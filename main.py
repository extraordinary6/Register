#!/usr/bin/env python
"""RegPulse CLI — generate Verilog register banks & UVM RAL models from Excel.

Usage:
    python main.py --input_excel spec.xlsx --output_dir ./output
    python main.py --input_excel spec.xlsx --output_dir ./output --format rtl,uvm,c_header
    python main.py --input_excel spec.xlsx --output_dir ./output --rtl_only
    python main.py --input_excel spec.xlsx --output_dir ./output --template_excel
"""

import argparse
import logging
import os
import sys
import traceback

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from src.parser.excel_parser import ExcelParser
from src.validators.validator import Validator, ValidationError
from src.generators.rtl_generator import RtlGenerator
from src.generators.apb_wrapper_generator import ApbWrapperGenerator
from src.generators.ahb_wrapper_generator import AhbWrapperGenerator
from src.generators.axi_wrapper_generator import AxiWrapperGenerator
from src.generators.uvm_generator import UvmGenerator
from src.generators.c_header_generator import CHeaderGenerator
from src.generators.json_generator import JsonGenerator
from src.generators.html_generator import HtmlGenerator
from src.generators.markdown_generator import MarkdownGenerator

logger = logging.getLogger("regpulse")

__version__ = "0.6.0"

ALL_FORMATS = ["rtl", "uvm", "c_header", "json", "html", "markdown"]
DEFAULT_FORMATS = ["rtl", "uvm", "c_header", "json", "html", "markdown"]

BUS_CHOICES = ["none", "apb", "ahb", "axi"]


def setup_logging(verbose: bool = False):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def main():
    parser = argparse.ArgumentParser(
        description="RegPulse — Generate APB Verilog, UVM RAL, C headers, JSON, and HTML from Excel register specs."
    )
    parser.add_argument("--input_excel", required=True,
                        help="Path to the input .xlsx register specification file.")
    parser.add_argument("--output_dir", required=True,
                        help="Directory where generated files will be written.")
    parser.add_argument("--block_name", default="reg_top",
                        help="Name for the register block / module (default: reg_top).")
    parser.add_argument("--base_address", default="0x0",
                        help="Base address for the register map (default: 0x0).")
    parser.add_argument("--data_width", type=int, default=32,
                        help="Register data width in bits (default: 32).")
    parser.add_argument("--bus", default="none", choices=BUS_CHOICES,
                        help="Bus protocol wrapper to generate: none, apb, ahb, axi (default: none).")
    parser.add_argument("--format", default=None,
                        help=f"Comma-separated list of output formats: {ALL_FORMATS}. "
                             f"Default: all.")
    parser.add_argument("--rtl_only", action="store_true",
                        help="Generate only Verilog RTL (shorthand for --format rtl).")
    parser.add_argument("--uvm_only", action="store_true",
                        help="Generate only UVM RAL (shorthand for --format uvm).")
    parser.add_argument("--template_excel", action="store_true",
                        help="Generate a blank Excel template file and exit.")
    parser.add_argument("--dry_run", action="store_true",
                        help="Parse and validate only; no file output.")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Enable verbose (DEBUG-level) logging output.")
    parser.add_argument("--version", action="version", version=f"RegPulse {__version__}")
    args = parser.parse_args()

    setup_logging(args.verbose)

    if args.rtl_only and args.uvm_only:
        parser.error("--rtl_only and --uvm_only are mutually exclusive.")

    # Resolve output formats
    if args.rtl_only:
        formats = ["rtl"]
    elif args.uvm_only:
        formats = ["uvm"]
    elif args.format:
        formats = [f.strip() for f in args.format.split(",")]
    else:
        formats = list(DEFAULT_FORMATS)

    for f in formats:
        if f not in ALL_FORMATS:
            parser.error(f"Unknown format '{f}'. Supported: {ALL_FORMATS}")

    # Generate template Excel if requested
    if args.template_excel:
        _generate_template(args.output_dir)
        return

    if not os.path.isfile(args.input_excel):
        logger.error("Input file not found: %s", args.input_excel)
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)

    # 1. Parse
    logger.info("Parsing Excel: %s", args.input_excel)
    excel_parser = ExcelParser(args.input_excel, block_name=args.block_name)
    bank = excel_parser.parse()
    logger.info("Parsed %d register(s), address space = %d bytes.",
                bank.num_registers, bank.address_space)

    # 2. Validate
    logger.info("Validating register definitions ...")
    validator = Validator(bank)
    validator.validate()
    logger.info("Validation passed.")

    # Dry-run: stop after validation
    if args.dry_run:
        logger.info("Dry run — no files generated.")
        return

    # 3. Generate outputs
    for fmt in formats:
        if fmt == "rtl":
            logger.info("Generating RTL (Verilog) ...")
            gen = RtlGenerator(bank)
            path = gen.generate(args.output_dir)
            logger.info("Written: %s", path)
        elif fmt == "uvm":
            logger.info("Generating UVM RAL (SystemVerilog) ...")
            gen = UvmGenerator(bank)
            path = gen.generate(args.output_dir)
            logger.info("Written: %s", path)
        elif fmt == "c_header":
            logger.info("Generating C header ...")
            gen = CHeaderGenerator(bank)
            path = gen.generate(args.output_dir)
            logger.info("Written: %s", path)
        elif fmt == "json":
            logger.info("Generating JSON IR ...")
            gen = JsonGenerator(bank)
            path = gen.generate(args.output_dir)
            logger.info("Written: %s", path)
        elif fmt == "html":
            logger.info("Generating HTML documentation ...")
            gen = HtmlGenerator(bank)
            path = gen.generate(args.output_dir)
            logger.info("Written: %s", path)
        elif fmt == "markdown":
            logger.info("Generating Markdown documentation ...")
            gen = MarkdownGenerator(bank)
            path = gen.generate(args.output_dir)
            logger.info("Written: %s", path)

    # 4. Generate bus wrapper if requested
    if args.bus != "none":
        bus = args.bus
        if bus == "apb":
            logger.info("Generating APB4 wrapper ...")
            wgen = ApbWrapperGenerator(bank)
        elif bus == "ahb":
            logger.info("Generating AHB-Lite wrapper ...")
            wgen = AhbWrapperGenerator(bank)
        elif bus == "axi":
            logger.info("Generating AXI4-Lite wrapper ...")
            wgen = AxiWrapperGenerator(bank)
        path = wgen.generate(args.output_dir)
        logger.info("Written: %s", path)

    logger.info("Done.")


def _generate_template(output_dir: str):
    """Generate a blank Excel template with data validation."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "register_template.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Registers"

    headers = ["Name", "Offset", "Field", "Bits", "Access", "Reset",
               "Hardware Trigger", "Side Effect"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Data validation for Access column
    dv_access = DataValidation(
        type="list",
        formula1='"RW,RO,W1C,RC,RS"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Access Type",
        error="Must be one of: RW, RO, W1C, RC, RS"
    )
    dv_access.add(f"E2:E1000")
    ws.add_data_validation(dv_access)

    # Data validation for Hardware Trigger column
    dv_hw = DataValidation(
        type="list",
        formula1='"input,output,"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid HW Interface",
        error='Must be "input", "output", or left blank.'
    )
    dv_hw.add(f"G2:G1000")
    ws.add_data_validation(dv_hw)

    # Set column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 24

    # Add a sample row
    ws.append(["MY_REG", "0x00", "ENABLE", "0", "RW", "0", "output", ""])

    wb.save(out_path)
    logger.info("Template written: %s", out_path)


if __name__ == "__main__":
    try:
        main()
    except ValueError as exc:
        logger.error("Parse error: %s", exc)
        sys.exit(1)
    except ValidationError as exc:
        logger.error("Validation error: %s", exc)
        sys.exit(2)
    except FileNotFoundError as exc:
        logger.error("File not found: %s", exc)
        sys.exit(3)
    except Exception as exc:
        logger.critical("Unexpected error: %s", exc)
        traceback.print_exc()
        sys.exit(99)
